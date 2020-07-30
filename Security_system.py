import cv2
import numpy as np
import face_recognition
import os
from datetime import date
from datetime import datetime
import openpyxl as xl
import mysql.connector
from pyzbar.pyzbar import decode
from pathlib import Path
import xlrd

vid_w = 1920
vid_h = 1080
row123 =1
count = 0

path = 'images'
images = []
classnames = []

myList = os.listdir(path)
print(myList)
for cl in myList:
    curImg = cv2.imread(f'{path}/{cl}')
    images.append(curImg)
    classnames.append(os.path.splitext(cl)[0])

print(classnames)

def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList

encodeListKnown = findEncodings(images)
print('encoding complete')

cap = cv2.VideoCapture(0)
cap.set(3,vid_w)
cap.set(4,vid_h)
while True:
    success,img = cap.read()
    imgS = cv2.resize(img,(0,0),None,0.25,0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    facesCurFrame = face_recognition.face_locations(imgS)
    encodesCureFrame = face_recognition.face_encodings(imgS, facesCurFrame)

    #print(cap.get(cv2.CAP_PROP_POS_FRAMES))

#    count = count + 1



    for encodeFace,faceLoc in zip(encodesCureFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDist = face_recognition.face_distance(encodeListKnown, encodeFace)
        print(faceDist)
        matchIndex = np.argmin(faceDist)

        if matches[matchIndex]:
            name = classnames[matchIndex]
            print(name)
            y1,x2,y2,x1 = faceLoc
            y1, x2, y2, x1 = y1*4,x2*4,y2*4,x1*4
            cv2.rectangle(img, (x1, y1), (x2, y2), (0,255,0),2)
            cv2.rectangle(img, (x1,y2-35), (x2,y2),(0,255,0),cv2.FILLED)
            cv2.putText(img, name, (x1 +6, y2-6), cv2.FONT_HERSHEY_SIMPLEX,1,(255,255,255),2)

            connection = mysql.connector.connect(host='localhost',
                                                 database='Students',
                                                 user='root',
                                                 password='root',
                                                 use_pure=True)



            if connection.is_connected():
                cursor = connection.cursor()

                cursor.execute(F"select stud_name from students where stud_name like '{name}'")
                confirm = cursor.fetchall()

                if len(confirm) > 0:
                    for barcode in decode(img):
                        mydata = barcode.data.decode('utf-8')
                        print(mydata)
                        pts = np.array([barcode.polygon], np.int32)
                        pts = pts.reshape(-1, 1, 2)
                        cv2.polylines(img, [pts], True, (255, 0, 0), 3)
                        pts2 = barcode.rect
                        cv2.putText(img, mydata, (pts2[0], pts2[1] - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 2)

                        cursor.execute("select * from students where reg= '%s'" % (mydata))
                        details = cursor.fetchall()
                        #x = details.split(", ")
                        print(details)


                        if len(details) > 0 and details[0][0] == name:
                            for row in details:
                                name123 = row[0]
                                reg = row[1]
                                marks = row[2]
                                gender = row[3]
                                age = row[4]

                                font_scale = 0.85
                                font = cv2.FONT_HERSHEY_SIMPLEX
                                text = "Access Granted: Welcome %s" % (name123)  # Count People at Risk
                                (text_width, text_height) = \
                                cv2.getTextSize(text, font, fontScale=font_scale, thickness=1)[0]
                                box_coords = ((10, 30), (vid_w, 25 - text_height - 10))
                                location = (10, 25)  # Set the location of the displayed text
                                cv2.rectangle(img, box_coords[0], box_coords[1], (0, 0, 0), cv2.FILLED)
                                cv2.putText(img, text, location, cv2.FONT_HERSHEY_SIMPLEX, font_scale, (255, 10, 10), 2,
                                            cv2.LINE_AA)

                                today = date.today()
                                d1 = today.strftime("%d/%m/%Y")
                                d1 = d1.translate({ord('/'): None})
                                now = datetime.now()
                                current_time = now.strftime("%H:%M")
                                my_file = Path(F"C:/Users/Gautam/Desktop/QR code scanner/entry logs/LOG_{d1}.xlsx")


                                if my_file.is_file() == False:
                                    path = "C:/Users/Gautam/Desktop/QR code scanner/entry logs"
                                    filename = F"LOG_{d1}.xlsx"
                                    wb = xl.Workbook()
                                    worksheet_1 = wb.active
                                    worksheet_1.title = "LOGS"
                                    wb.save(F'C:/Users/Gautam/Desktop/QR code scanner/entry logs/LOG_{d1}.xlsx')
                                    os.chmod(F'C:/Users/Gautam/Desktop/QR code scanner/entry logs/LOG_{d1}.xlsx', 0o777)
                                else:
                                    o = 1

                                loc12345 = (F"C:/Users/Gautam/Desktop/QR code scanner/entry logs/LOG_{d1}.xlsx")

                                # To open Workbook
                                book = xlrd.open_workbook(loc12345)
                                sheet = book.sheet_by_index(0)
                                # test = sheet.row_values(1)
                                row_count = sheet.nrows
                                print(row_count)
                                # test1 = sheet.row_values(2)
                                book.release_resources()

                                os.chmod(F'C:/Users/Gautam/Desktop/QR code scanner/entry logs/LOG_{d1}.xlsx', 0o777)
                                path1 = F"C:/Users/Gautam/Desktop/QR code scanner/entry logs/LOG_{d1}.xlsx"
                                wb = xl.load_workbook(path1)
                                worksheet_1 = wb["LOGS"]

                                c1 = worksheet_1.cell(row=1, column=1)
                                c1.value = "NAME"

                                c2 = worksheet_1.cell(row=1, column=2)
                                c2.value = "Registration no."

                                c3 = worksheet_1.cell(row=1, column=3)
                                c3.value = "Gender"

                                c4 = worksheet_1.cell(row=1, column=4)
                                c4.value = "Age"

                                c5 = worksheet_1.cell(row=1, column=5)
                                c5.value = "Time"
                                list = [name123, reg, gender, age, current_time]

                                if count >= 1:


                                    loc12345 = (F"C:/Users/Gautam/Desktop/QR code scanner/entry logs/LOG_{d1}.xlsx")

                                    # To open Workbook
                                    book = xlrd.open_workbook(loc12345)
                                    sheet = book.sheet_by_index(0)
                                    test = sheet.row_values(row_count - 1)
                                    row_count = sheet.nrows
                                    print(row_count)
                                    #test1 = sheet.row_values(2)
                                    book.release_resources()


                                count = count +1

                                if count >= 2:
                                    if list != test:
                                        row_count = row_count + 1
                                        col = 1
                                        for i in range(len(list)):
                                            col = i + 1
                                            wcell1 = worksheet_1.cell(row_count, col)
                                            wcell1.value = list[i]
                                    wb.save(F'C:/Users/Gautam/Desktop/QR code scanner/entry logs/LOG_{d1}.xlsx')
                                    wb.close()

                                else:
                                    col = 1
                                    row_count = row_count +1
                                    for i in range(len(list)):
                                        col = i + 1
                                        wcell1 = worksheet_1.cell(row_count, col)
                                        wcell1.value = list[i]
                                wb.save(F'C:/Users/Gautam/Desktop/QR code scanner/entry logs/LOG_{d1}.xlsx')
                                wb.close()
                                #time.sleep(1)

                        else:
                            font_scale = 0.85
                            font = cv2.FONT_HERSHEY_SIMPLEX
                            text = "Access Denied "  # Count People at Risk
                            (text_width, text_height) = \
                            cv2.getTextSize(text, font, fontScale=font_scale, thickness=1)[0]
                            box_coords = ((10, 30), (vid_w, 25 - text_height - 10))
                            location = (10, 25)  # Set the location of the displayed text
                            cv2.rectangle(img, box_coords[0], box_coords[1], (0, 0, 0), cv2.FILLED)
                            cv2.putText(img, text, location, cv2.FONT_HERSHEY_SIMPLEX, font_scale, (255, 10, 10), 2,
                                        cv2.LINE_AA)




    cv2.imshow('webcam', img)
    cv2.waitKey(1)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

